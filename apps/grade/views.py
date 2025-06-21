from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from django.http import HttpResponse
from .models import GradeItem, FinalGrade
from account.models import Student
from classroom.models import Class, Enrollment
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from config.decorators import role_required
from collections import defaultdict

# Create your views here.

# # # # # # # # # # # # #
# STUDENT 
# # # # # # # # # # # # #

def focus_starplot():
    # From student dashboard, focus the starplot window in the center to show the main starplot
    return None


def view_subject_grade():
    # Click the subject from starplot to show the percentage of grades
    return None



# # # # # # # # # # # # #
# TEACHER 
# # # # # # # # # # # # #


def __parse_file(file):
    wb = load_workbook(file)
    sheet = wb.active

    # Format: "Quarter 1 | Class: 9-Sapphire | Subject: Mathematics | Teacher: Mr. Reyes"
    header = sheet["A1"].value
    if not header or "Quarter" not in header or "Class:" not in header:
        raise ValueError("Missing header info in A1")

    parts = header.split("|")
    quarter = int(parts[0].strip().split(" ")[1])
    class_name = parts[1].split(":")[1].strip()
    subject = parts[2].split(":")[1].strip()
    teacher_name = parts[3].split(":")[1].strip()

    # Process students from row 5 downward
    students_data = []
    for row in sheet.iter_rows(min_row=5, values_only=True):
        name, lrn = row[0], str(row[1]).strip() if row[1] else None
        if not lrn:
            continue

        written = [(row[i] or 0, sheet.cell(row=4, column=i+1).value or 0) for i in range(2, 8)]
        performance = [(row[i] or 0, sheet.cell(row=4, column=i+1).value or 0) for i in range(8, 12)]
        assessment = [(row[i] or 0, sheet.cell(row=4, column=i+1).value or 0) for i in [12]]

        students_data.append({
            'name': name,
            'lrn': lrn,
            'written': written,
            'performance': performance,
            'assessment': assessment,
        })

    return class_name, subject, teacher_name, quarter, students_data

# # # # # # # # # # # # # # # # # # # # # # # # #


def __compute_grade(student, class_obj, quarter):
    # Fetch grading scheme and grade items
    scheme = class_obj.grading_scheme
    items = GradeItem.objects.filter(student=student, class_obj=class_obj, quarter=quarter)

    total_scores = {'WW': 0, 'PT': 0, 'QA': 0}
    total_max = {'WW': 0, 'PT': 0, 'QA': 0}

    for item in items:
        total_scores[item.component] += item.score
        total_max[item.component] += item.highest_possible_score

    def weighted(component, weight):
        if total_max[component] == 0:
            return 0
        percent = (total_scores[component] / total_max[component]) * 100
        return percent * weight

    final_grade = sum([
        weighted('WW', scheme.written_work_weight),
        weighted('PT', scheme.performance_task_weight),
        weighted('QA', scheme.quarterly_assessment_weight),
    ])

    final_grade = round(final_grade)

    # Save to FinalGrade model
    grade_obj, _ = FinalGrade.objects.get_or_create(
        student=student,
        class_obj=class_obj,
        quarter=quarter
    )
    grade_obj.final_grade = final_grade
    grade_obj.save()

    return final_grade


# # # # # # # # # # # # # # # # # # # # # # # # #


def __post_grade(students_data, class_obj, quarter):
    for data in students_data:
        try:
            student = Student.objects.get(lrn=data['lrn'])
        except Student.DoesNotExist:
            continue  # Skip unknown LRN

        # Delete old entries for this quarter
        GradeItem.objects.filter(student=student, class_obj=class_obj, quarter=quarter).delete()

        # Create new GradeItems
        for score, max_score in data['written']:
            if max_score == 0:
                continue
            GradeItem.objects.create(
                student=student,
                class_obj=class_obj,
                quarter=quarter,
                component='WW',
                score=score,
                highest_possible_score=max_score
            )

        for score, max_score in data['performance']:
            if max_score == 0:
                continue
            GradeItem.objects.create(
                student=student,
                class_obj=class_obj,
                quarter=quarter,
                component='PT',
                score=score,
                highest_possible_score=max_score
            )

        for score, max_score in data['assessment']:
            if max_score == 0:
                continue
            GradeItem.objects.create(
                student=student,
                class_obj=class_obj,
                quarter=quarter,
                component='QA',
                score=score,
                highest_possible_score=max_score
            )

        # Compute and save final grade
        __compute_grade(student, class_obj, quarter)

# # # # # # # # # # # # # # # # # # # # # # # # #

@login_required
@role_required('teacher')
def download_file(request, class_id):
    # Dowload the template Excel sheet for the teachers to use
    class_obj = Class.objects.get(id=class_id)
    students = Enrollment.objects.filter(class_obj=class_obj).select_related('student')

    wb = Workbook()
    ws = wb.active
    ws.title = "Quarter Template"

    # Header row (merged)
    header = f"Quarter 1 | Class: {class_obj.class_name} | Subject: {class_obj.subject} | Teacher: {class_obj.teacher}"
    ws.merge_cells('A1:M1')
    ws['A1'] = header

    # Category headers (row 2)
    ws.merge_cells('C2:H2')  # WW1–WW6
    ws.merge_cells('I2:L2')  # PT1–PT4
    ws.merge_cells('M2:M2')  # QA1
    ws['C2'] = 'Written Works'
    ws['I2'] = 'Performance Tasks'
    ws['M2'] = 'Quarterly Assessment'

    # Label headers (row 3)
    headers = ['Name', 'LRN'] + [f'WW{i+1}' for i in range(6)] + [f'PT{i+1}' for i in range(4)] + ['QA1']
    ws.append(headers)

    # Max score input row (row 4)
    ws.append(['', 'MAX SCORE'] + ['' for _ in range(11)])

    # Student rows
    for enrollment in students:
        student = enrollment.student
        ws.append([student.user.get_full_name(), student.lrn] + ['' for _ in range(11)])

    # Adjust column widths
    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.row_dimensions[1].height = 25
    
    # Set column widths for better readability
    ws.column_dimensions['A'].width = 30  # Full Name
    ws.column_dimensions['B'].width = 20  # LRN

    # Center alignment for merged cells
    for cell in ['C2', 'I2', 'M2']:
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')

    # Download response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="grade_template_{class_obj.class_name}.xlsx"'
    wb.save(response)
    return response

# # # # # # # # # # # # # # # # # # # # # # # # #

@login_required
@role_required('teacher')
def upload_grade(request, class_id):
    class_obj = get_object_or_404(Class, id=class_id)
    enrolled_students = list(class_obj.enrollments.select_related('student').all())
    enrolled_lrns = set([en.student.lrn for en in enrolled_students])

    if request.method == 'POST':
        try:
            quarter = int(request.POST.get('quarter'))
            if quarter not in [1, 2, 3, 4]:
                raise ValueError("Quarter must be 1 to 4.")
        except (ValueError, TypeError):
            messages.error(request, "Invalid quarter specified.")
            return redirect('upload_grade', class_id=class_id)

        file = request.FILES.get('grades_file')
        if not file:
            messages.error(request, "Please upload a valid Excel file.")
            return redirect('upload_grade', class_id=class_id)

        try:
            uploaded_class_name, uploaded_subject, teacher_name, quarter_from_file, students_data = __parse_file(file)


        except Exception as e:
            messages.error(request, f"Excel parsing failed: {e}")
            return redirect('upload_grade', class_id=class_id)

        # ✅ Validate class name and subject
        if uploaded_class_name.lower() != class_obj.class_name.lower():
            messages.error(request, f"Class name mismatch! Excel says: '{uploaded_class_name}' — expected: '{class_obj.class_name}'")
            return redirect('upload_grade', class_id=class_id)

        if uploaded_subject.lower() != class_obj.subject.lower():
            messages.error(request, f"Subject mismatch! Excel says: '{uploaded_subject}' — expected: '{class_obj.subject}'")
            return redirect('upload_grade', class_id=class_id)
        
        if quarter != quarter_from_file:
            messages.error(
        request,
        f"Mismatch: You selected Quarter {quarter}, but the uploaded Excel file says Quarter {quarter_from_file}."
        )
            return redirect('upload_grade', class_id=class_id)


        # ✅ Validate LRN completeness
        uploaded_lrns = set(s['lrn'] for s in students_data)
        missing_lrns = enrolled_lrns - uploaded_lrns
        unknown_lrns = uploaded_lrns - enrolled_lrns

        if missing_lrns:
            messages.error(request, f"Missing LRNs: {', '.join(missing_lrns)}")
            return redirect('upload_grade', class_id=class_id)

        if unknown_lrns:
            messages.error(request, f"Unrecognized LRNs: {', '.join(unknown_lrns)} (not enrolled in this class)")
            return redirect('upload_grade', class_id=class_id)

        # ✅ Post grades
        __post_grade(students_data, class_obj, quarter)
        messages.success(request, f"Grades for Quarter {quarter} successfully uploaded.")

        messages.success(request, f"Grades for Quarter {quarter_from_file} successfully uploaded.")
        return redirect('view_class_list', class_id=class_id)

    return render(request, 'grade/teacher/upload_grade.html', {'class_obj': class_obj})




from django.shortcuts import render, get_object_or_404
from account.models import Student
from classroom.models import Class
from .models import GradeItem

# # # # # # # # # # # # # # # # # # # # # # # # #


@login_required
@role_required('teacher')
def view_grade(request, student_id, class_id, quarter):
    student = get_object_or_404(Student, id=student_id)
    class_obj = get_object_or_404(Class, id=class_id)
    grading_scheme = class_obj.grading_scheme

    def compute_component(items):
        total_score = sum(item.score for item in items)
        total_max = sum(item.highest_possible_score for item in items)
        return (total_score / total_max) * 100 if total_max else 0

    grades = {}
    for q in range(1, 5):
        items = GradeItem.objects.filter(student=student, class_obj=class_obj, quarter=q)
        ww = items.filter(component='WW')
        pt = items.filter(component='PT')
        qa = items.filter(component='QA')

        final_grade = round(
            compute_component(ww) * grading_scheme.written_work_weight +
            compute_component(pt) * grading_scheme.performance_task_weight +
            compute_component(qa) * grading_scheme.quarterly_assessment_weight
        ) if items.exists() else None

        grades[q] = {
            'written_work': ww,
            'performance_task': pt,
            'quarterly_assessment': qa,
            'final_grade': final_grade,
        }

    context = {
        'student': student,
        'class_obj': class_obj,
        'grades': grades,
        'quarters': range(1, 5),
        'selected_quarter': quarter,
    }

    return render(request, 'grade/teacher/view_grade.html', context)



# # # # # # # # # # # # # # # # # # # # # # # # #

@login_required
@role_required('teacher')
def edit_grade():
    # Edit the grade of an individual student
    return None
